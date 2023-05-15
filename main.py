import openpyxl
import requests
import json
from openpyxl import workbook, load_workbook



# User edited parameters:

month = "H" #column name of month we are sending for
due_date = "30th Of May" #a sentence that informs the reader when the bill is due by.



# declaring all constants and classes to work through for remaining file

wb = load_workbook("Student & Revenue Sheet.xlsx")
ws = wb['Student Master Sheet']
row_count = ws.max_row
sheets = wb.sheetnames

url = "(Enter URL here)"

class ColNamesMismatch(Exception):
    pass

headers = {
    "Content-Type": "application/json",
    #"Authorization": "Bearer (Enter authentication token here)"
}


#Function defination begins.
def send_msg(phone_num, name, amt_owed):
    # send message with given parameters

    data = {
        "messaging_product": "whatsapp",
        "to": str(phone_num),
        "type": "template",
        "template": { "name": "reminder", "language": { "code": "en_US" },
                      "components": [{
                          "type": "body",
                          "parameters": [{
                              "type": "text",
                              "text": name
                          },
                              {
                                  "type": "text",
                                  "text": str(amt_owed)
                              },
                              {
                                  "type": "text",
                                  "text": due_date
                              }
                          ]
                      }]
                      }
    }

    response = requests.post(url, headers=headers, data=json.dumps(data))
    print(response.json())

    return response.json()


def amt_owed(stud_name, total_owed):
    # Simple function to calculate the amount owed by given user.
    return total_owed - total_paid_stud(stud_name)


def total_paid_stud(stud_name):
    # finds the total amount paid by a given student across all sheets for the selected month.

    tot_amt = 0

    for sheet in sheets[1:]:
        ws_local = wb[sheet]

        for row in range(2, ws_local.max_row+1):
            if stud_name == ws_local['A' + str(row)].value:
                amt = ws_local[month + str(row)].value
                if amt == None:
                    tot_amt += 0
                else:
                    tot_amt += amt



    return tot_amt


def send_check(amt):
    # Returns status code of message depending on due amount.
    if amt == 0:
        return "N"
    elif amt < 0:
        return "E"
    else:
        return "Y"

def send_allmsgs():
    # Traverses file and sends all messages. In a function to help with testing but could be done globally.

    workbook_temp = openpyxl.Workbook()
    sheet_temp = workbook_temp.active

    for row in range(2, row_count+1):

        if not(ws["A1"].value == "Name") or not(ws["E1"].value == "Mobile number") or not(ws["F1"].value == "Amount due"):
            raise ColNamesMismatch("The column names do not match with previous indexes. Have the Colums been renamed/rearranged?")

        name = ws['A' + str(row)].value
        amt_due = ws['F' + str(row)].value
        tel_num = ws['E' + str(row)].value

        #print(name, amt_due, tel_num) uncomment to debug

        amt = amt_owed(name, amt_due)
        stat_code = send_check(amt)

        if stat_code == "Y":
            send_msg(tel_num, name, amt)
            sheet_temp["A" + str(row)] = name
            sheet_temp["B" + str(row)] = "Y"

        elif stat_code == "E":
            sheet_temp["A" + str(row)] = name
            sheet_temp["B" + str(row)] = "E"

        else:
            sheet_temp["A" + str(row)] = name
            sheet_temp["B" + str(row)] = "N"

    workbook_temp.save('messages_log.xlsx')

    return None

# function call to start sending the messages.
